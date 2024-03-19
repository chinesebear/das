import os

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import pandas as pd
import numpy as np

def get_ctnr_num(input_path):
    raw_data = input_path.split('-')
    file_name = raw_data[-1].split('.')
    ctnr_number = file_name[0]
    return ctnr_number.strip()

def get_short_invoice(invoice):
    raw_data = invoice.split('-')
    short_invoice = f"{raw_data[1]}-{raw_data[2]}"
    return short_invoice

def create_new_excel(path):
    if os.path.exists(path):
        return
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Draft DAS"
    wb.save(path)

def set_sheet_col_style(sheet):
    sheet.column_dimensions['A'].width=20
    sheet.column_dimensions['B'].width=10
    sheet.column_dimensions['C'].width=20
    sheet.column_dimensions['D'].width=10
    sheet.column_dimensions['E'].width=30
    sheet.column_dimensions['F'].width=20
    sheet.column_dimensions['G'].width=20
    sheet.column_dimensions['H'].width=20
    max_rows = sheet.max_row  # 获取最大行
    max_columns = sheet.max_column  # 获取最大列
    for i in range(1, max_rows + 1):
        for j in range(1, max_columns + 1):
            sheet.cell(i, j).alignment = Alignment(horizontal='center',vertical='center')    
    
def excel_process(input_path, output_path, invoice):
    wb = load_workbook(input_path)
    sheets = wb.worksheets   # 获取当前所有的sheet
    print(sheets)

    # 获取第一张sheet
    sheet = sheets[0]
    print(sheet)
    short_invoice = get_short_invoice(invoice)
    ctnr_number = get_ctnr_num(input_path)
    company = sheet['A1'].value
    report_name = sheet['A2'].value
    module_type = sheet['B4'].value
    customer = sheet['E4'].value
    date_val = sheet['J4'].value

    data = {
    "Voc": [sheet['B5'].value, sheet['B6'].value, sheet['B7'].value],
    "Isc": [sheet['C5'].value, sheet['C6'].value, sheet['C7'].value],
    "Vpm": [sheet['D5'].value, sheet['D6'].value, sheet['D7'].value],
    "Ipm": [sheet['E5'].value, sheet['E6'].value, sheet['E7'].value],
    "Pm": [sheet['F5'].value, sheet['F6'].value, sheet['F7'].value],
    "FF":[sheet['G5'].value, sheet['G6'].value, sheet['G7'].value],
    }
    df_stat = pd.DataFrame(data)

    max_row = sheet.max_row
    max_col = sheet.max_column
    print(f"max_row:{max_row}, max_col:{max_col}")

    df = pd.DataFrame(columns=['Module_ID', 'Voc', 'Isc', 'Vpm','Ipm', 'Pm', 'FF', 'Watt_Marking', 'Pallet_ID', 'Current_level', 'Shipment_date', 'License_plate_number'])

    for i in range(9, max_row+1):
        row_data = []
        for j in range(1, max_col+1):
            val = sheet.cell(i,j).value
            if val is None: 
                break 
            row_data.append(sheet.cell(i,j).value)
        if row_data == []:
            break
        df.loc[i-9]=row_data
    print(df)
    Pallet_ID_counts = df['Pallet_ID'].value_counts()
    print(Pallet_ID_counts)

    if not os.path.exists(output_path):
        create_new_excel(output_path)
    wb_out = load_workbook(output_path)
    sheet = wb_out.active
    head = ['PRODUCT','PCS','PALLET ID','LOCATION','INVOICE/CTNR_NUMBER','WAREHOUSE REF IN','CONSIGNEE','ORDER NO']
    tail = ['','','','','','','','']
    sheet.append(head)
    for id, count in Pallet_ID_counts.items():
        PRODUCT = module_type
        PCS = count
        PALLET_ID = id
        LOCATION = 'RWB'
        INVOICE_CTNR_NUMBER = f'{short_invoice}/{ctnr_number}'
        WAREHOUSE_REF_IN = 'FREE'
        CONSIGNEE = 'DASSOLFRA'
        ORDER_NO = invoice
        data = [PRODUCT, PCS, PALLET_ID, LOCATION, INVOICE_CTNR_NUMBER, WAREHOUSE_REF_IN, CONSIGNEE, ORDER_NO]
        sheet.append(data)
    sheet.append(tail)
    set_sheet_col_style(sheet)
    wb_out.save(output_path)
    
if __name__=="__main__":
    input_path = 'tmpw6er8scq/Flash Report-720pcs-600W-2024-2-27 - CSNU7329467.xlsx'
    output_path = 'tmpw6er8scq/result.xlsx'
    invoice = 'DASBH-N15-240104'
    excel_process(input_path, output_path, invoice)
    # wb = load_workbook('/home/yang/sda/github/das/output/Receive Upload_BL No.COSU1710812845.xlsx')
    # sheet = wb.active   # 获取当前所有的sheet
    # set_sheet_col_style(sheet)
    # wb.save('/home/yang/sda/github/das/output/Receive Upload_BL No.COSU1710812845.xlsx')